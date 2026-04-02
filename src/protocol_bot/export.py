import pandas as pd
import bisect
from openpyxl.styles import Border, Side, Font
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from pathlib import Path

class FileGeneration:
    def __init__(self):
        pass
    
    def composite_file(self, result, stock_repo, path_config, author=None, time=None, title=None, time_break=None):
        # Ensure we are working with a Path object
        comp_result = result.comp
        well_names = result.well
        protocol_summary = result.protocol
        output_path = Path(path_config.output_file)
        
        # Create the directory if it doesn't exist
        output_path.parent.mkdir(parents=True, exist_ok=True)

        comp_result = self.add_headline(author, time, title, comp_result, time_break)
        self.min_volume_check(comp_result)
        self.generate_excel_file(comp_result, well_names, protocol_summary, stock_repo, output_path)

    def generate_excel_file(self, comp_result, well_names, protocol_summary, stock_repo, output_path):
        df = pd.DataFrame(comp_result)
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, header=False, sheet_name='Experimental_Protocol')
            worksheet = writer.sheets['Experimental_Protocol']
            number_font = Font(name="Courier New")
            
            # Set column widths using column letters
            worksheet.column_dimensions['A'].width = 20
            worksheet.column_dimensions['B'].width = 30
            worksheet.column_dimensions['C'].width = 30
            worksheet.column_dimensions['D'].width = 30
            worksheet.column_dimensions['E'].width = 20
            worksheet.column_dimensions['F'].width = 20
            worksheet.column_dimensions['G'].width = 20
            worksheet.column_dimensions['H'].width = 30
            worksheet.column_dimensions['I'].width = 30
            worksheet.column_dimensions['J'].width = 30
            worksheet.column_dimensions['K'].width = 20

            # --- Aesthetic design for organizing --- #
            # Add the border line for the excel file
            bold_border = Border(
                left=Side(border_style='thick'),
                right=Side(border_style='thick')
            )

            bold_border_bottom = Border(
                bottom=Side(border_style='thick')
            )

            last_row = worksheet.max_row
            last_col = worksheet.max_column

            for row in range(4, last_row + 1):
                init_cell = f"I{row}"
                final_cell = f"J{row}"
                for col in ['B', 'H']:
                    worksheet[f"{col}{row}"].border = bold_border

            for col in range(1, last_col+1):
                cell = worksheet.cell(row=3, column=col)
                cell.border = bold_border_bottom

            # Add the color patches for the excel file
            component_colors = {}
            palette = [
                "2986CC", "FFD966", "B99FBB", "E69138", "F44336", "16537E",
                "8FCE00", "7E96C4", "F7E4E0", "52796F", "6E6879", "BE8C00",
                "B7AA35", "E14832", "85A7AD", "608956", "74B354", "7EABCC",
                "97A4AF", "F1D6D6", "E99455", "E6E0CC", "FFFED6", "ACAB97",
                "B39726", "BFA7A7", 'F1F2E0', 'EFD8D8', '95BBE3', 'B0D8D8',
                'EDD3A4', '9396D5', '78F06F', 'F7AD4E', 'AABCCD', 'BDA996'
            ]

            color_index = 0
            prev_empty1 = True
            noter = False
            noter1 = False

            for row in worksheet.iter_rows(min_row=5, max_row=last_row):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cell.font = number_font
                cell1 = row[1]
                cell2 = row[7]

                val1 = str(cell1.value).strip() if cell1.value is not None else ""
                val2 = str(cell2.value).strip() if cell2.value is not None else ""

                if val1 != "" and val1 != "Milli Q H2O":
                    if val1 not in stock_repo:
                        if val1 not in component_colors:
                            color = palette[color_index]
                            color_index += 1
                            component_colors[val1] = color
                        else:
                            color = component_colors[val1]
                        cell1.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

                if val2 != "" and val2 != "Milli Q H2O":
                    if val2 not in stock_repo and val2 not in well_names:
                        if val2 not in component_colors:
                            color = palette[color_index % len(palette)]
                            color_index += 1
                            component_colors[val2] = color
                        else:
                            color = component_colors[val2]
                        cell2.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

                # Add the grey box for each block
                cell3 = row[4]
                val3 = str(cell3.value).strip() if cell3.value is not None else ""
                cell4 = row[10]
                val4 = str(cell4.value).strip() if cell4.value is not None else ""
                color_grey = "DDDDDD"

                if noter:
                    row[2].fill = PatternFill(start_color=color_grey, end_color=color_grey, fill_type="solid")
                    row[3].fill = PatternFill(start_color=color_grey, end_color=color_grey, fill_type="solid")
                    noter = False

                if val3 == "Volume Total":
                    row[2].fill = PatternFill(start_color=color_grey, end_color=color_grey, fill_type="solid")
                    row[3].fill = PatternFill(start_color=color_grey, end_color=color_grey, fill_type="solid")
                    noter = True
                
                if val1 == "Milli Q H2O":
                    row[2].fill = PatternFill(start_color=color_grey, end_color=color_grey, fill_type="solid")
                    row[3].fill = PatternFill(start_color=color_grey, end_color=color_grey, fill_type="solid")

                if noter1:
                    row[8].fill = PatternFill(start_color=color_grey, end_color=color_grey, fill_type="solid")
                    row[9].fill = PatternFill(start_color=color_grey, end_color=color_grey, fill_type="solid")
                    noter1 = False

                if val4 == "Volume Total":
                    row[8].fill = PatternFill(start_color=color_grey, end_color=color_grey, fill_type="solid")
                    row[9].fill = PatternFill(start_color=color_grey, end_color=color_grey, fill_type="solid")
                    noter1 = True
                
                if val2 == "Milli Q H2O":
                    row[8].fill = PatternFill(start_color=color_grey, end_color=color_grey, fill_type="solid")
                    row[9].fill = PatternFill(start_color=color_grey, end_color=color_grey, fill_type="solid")

            df2 = pd.DataFrame(
                protocol_summary,
                columns=["Solution Name", "Final Concentration", "Final Volume", "Date Created"]
            )
            df2.to_excel(writer, index=False, sheet_name="Protocol_Summary")
            ws2 = writer.sheets["Protocol_Summary"]
            for col in ws2.columns:
                max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
                ws2.column_dimensions[col[0].column_letter].width = max_len + 2
        print(f"✅ Protocol Excel generated successfully at: {output_path}")
    
    def add_headline(self, name, time, title, prev_result, time_break):
        caveat1 = "Units for initial and final concentrations are listed within the Material names."
        caveat2 = "If not present, units are in uM (micromolar)."
        if time_break is not None:
            note = "The time break of fluorescence measurements are set at: " + ", ".join([str(t) + " min" for t in time_break])
            headline = [[name, None, caveat1, None, None, None, None, note, None, None, None], 
                        [time, None, caveat2, None, None, None, None, None, None, None, None],
                        [title, None, None, None, None, None, None, None, None, None, None],
                        [None, 'Material', 'Initial Concentration', 'Final Concentration', 'Volume', None, None, 'Material', 'Initial Concentration', 'Final Concentration', 'Volume']]
        else:
            headline = [[name, None, caveat1, None, None, None, None, None, None, None, None], 
                        [time, None, caveat2, None, None, None, None, None, None, None, None],
                        [title, None, None, None, None, None, None, None, None, None, None],
                        [None, 'Material', 'Initial Concentration', 'Final Concentration', 'Volume', None, None, 'Material', 'Initial Concentration', 'Final Concentration', 'Volume']]
        headline.extend(prev_result)
        return headline
    
    def min_volume_check(self, comp_result):
        for i in range(4, len(comp_result)):
            row = comp_result[i]
            for j, cell in enumerate(row):
                    if isinstance(cell, (int, float)):
                        if j in [4, 10]:
                            cell = round(cell, 3)
                            if cell < 0.1:
                                cell = 0
                        else:
                            cell = round(cell, 5)
                        comp_result[i][j] = cell

        for i in range(4, len(comp_result)):
            row = comp_result[i]
            vol1 = row[4]
            vol2 = row[10]
            if vol1 is not None:
                if row[1] is not None:
                    if vol1 < 0.6 and vol1 != 0:
                        print(f"[WARNING]: The volume for '{row[1]}' in row {i+1} is below the minimum threshold: 0.6 uL")
                        print(f"Current volume: {vol1}")
                        print(f"Please adjust the conc for the protocol recipe")
            if vol2 is not None:
                if row[7] is not None:
                    if vol2 < 0.6 and vol2 != 0:
                        print(f"[WARNING]: The volume for '{row[7]}' in row {i+1} is below the minimum threshold: 0.6 uL")
                        print(f"Current volume: {vol2}")
                        print(f"Please adjust the conc for the protocol recipe")

    def formula_generation(self, worksheet, volume):
        # Add the formula portion to the well plate
        last_row = worksheet.max_row
        
        # Add the formula for volume calculation in the well plate
        for i, row in enumerate(worksheet.iter_rows(min_row=5, max_row=last_row), start=5):
            name = str(row[7].value).strip() if row[7].value is not None else ""
            if name != "Milli Q H2O":
                val1 = str(row[8].value).strip() if row[8].value is not None else ""
                val2 = str(row[9].value).strip() if row[9].value is not None else ""
                if val1 != "" and val2 != "":
                    col_H = get_column_letter(9)
                    col_I = get_column_letter(10)
                    formula = f"={col_I}{i}*{volume}/{col_H}{i}"
                    worksheet.cell(row=i, column=11).value = formula
        
        prev_indc = False
        indic = 5
        for i, row in enumerate(worksheet.iter_rows(min_row=5, max_row=last_row), start=5):
            name = str(row[7].value).strip() if row[7].value is not None else ""
            if prev_indc == True:
                if name != "":
                    indic = i
                    prev_indc = False
            
            if name == "Milli Q H2O":
                col = get_column_letter(11)
                formula = f"={col}{i+2}-SUM({col}{indic}:{col}{i-1})"
                worksheet.cell(row=i, column=11).value = formula
            
            if name == "":
                prev_indc = True
        
        # Add the formula for volume calculation in other place
        idx_list = []
        for i, row in enumerate(worksheet.iter_rows(min_row=5, max_row=last_row), start=5):
            name = str(row[4].value).strip() if row[4].value is not None else ""
            if name == "Volume Total":
                idx_list.append(i+1)

        prev_row = False
        for i, row in enumerate(worksheet.iter_rows(min_row=5, max_row=last_row), start=5):
            name = str(row[1].value).strip() if row[1].value is not None else ""
            if name != "Milli Q H2O":
                if row[3].value is not None:
                    if str(row[3].value).strip() != "Aliquot Volume":
                        if prev_row == False:
                            val1 = str(row[3].value).strip()
                        else:
                            val1 = ""
                            prev_row = False
                    else:
                        val1 = ""
                        prev_row = True
                else:
                    val1 = ""
                    prev_row = False

                if row[4].value is not None and str(row[4].value).strip() != "Volume Total":
                    val2 = str(row[4].value).strip()
                else:
                    val2 = ""
                if val1 != "" and val2 != "":
                    index = bisect.bisect_left(idx_list, i)
                    volume = worksheet.cell(row=idx_list[index], column=5).value
                    col_H = get_column_letter(3)
                    col_I = get_column_letter(4)
                    formula = f"={col_I}{i}*{get_column_letter(5)}{idx_list[index]}/{col_H}{i}"
                    worksheet.cell(row=i, column=5).value = formula
        
        prev_indc2 = True
        indic = 5
        for i, row in enumerate(worksheet.iter_rows(min_row=5, max_row=last_row), start=5):
            name = str(row[1].value).strip() if row[1].value is not None else ""
            val1 = str(row[2].value).strip() if row[2].value is not None else ""
            if prev_indc2 == True:
                if name != "" and val1 != "":
                    indic = i
                    prev_indc2 = False
            
            if name == "Milli Q H2O":
                col = get_column_letter(5)
                formula = f"={col}{i+2}-SUM({col}{indic}:{col}{i-1})"
                worksheet.cell(row=i, column=5).value = formula
            
            if name == "":
                prev_indc2 = True
   
        # Add the formula for total volume calculation in prepare the stock solution
        prev_indc1 = True
        store = None
        for i, row in enumerate(worksheet.iter_rows(min_row=5, max_row=last_row), start=5):
            formula = None
            name = str(row[1].value).strip() if row[1].value is not None else ""
            name1 = str(row[4].value).strip() if row[4].value is not None else ""
            name_list = []
            if name != "":
                if prev_indc1:
                    for i, row in enumerate(worksheet.iter_rows(min_row=5, max_row=last_row), start=5):
                        match1 = str(row[1].value).strip() if row[1].value is not None else ""
                        match2 = str(row[7].value).strip() if row[7].value is not None else ""
                        value1 = str(row[4].value).strip() if row[4].value is not None else ""
                        if value1 != "":
                            if match1 == name:
                                name_list.append((get_column_letter(5), i))
                        if match2 == name:
                            name_list.append((get_column_letter(11), i))
            if len(name_list) != 0:
                formula = f"=({'+'.join([f'{col}{row}' for col, row in name_list])})*1.2"
                store = formula
            if name1 == "Volume Total":
                worksheet.cell(row=i+1, column=5).value = store
            if name == "":
                prev_indc1 = True
                formula = None
            else:
                prev_indc1 = False
        
        return worksheet