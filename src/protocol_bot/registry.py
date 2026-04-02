from openpyxl import Workbook, load_workbook
from datetime import date
from pathlib import Path

class TrackRegistry:
    def __init__(self):
        pass

    def updateCentralRegistry(self, repo_SPIs, registry_path, stock):
        used_protocols = set()
        for spi in repo_SPIs:
            if spi.isBulk or spi.isPrebulk or spi.well is not None:
                continue
            used_protocols.add(spi.SProtocol.Name)

        registry_path = Path(registry_path)

        if not registry_path.exists():
            wb = Workbook()
            ws = wb.active
            ws.title = "Solution_Registry"
            ws.append([
                "Date Created", 
                "Solution Name", 
                "Final Concentration",
                "Final Volume"
            ])
            wb.save(registry_path)

        wb = load_workbook(registry_path)
        ws = wb["Solution_Registry"]

        existing_names = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            existing_names.add(row[1])

        for spi in repo_SPIs:
            sp = spi.SProtocol
            for component in sp.Components:
                if component in stock and component not in existing_names:
                    ws.append([
                        date.today().isoformat(),
                        component,
                        stock[component],
                        "N/A"
                    ])
                    existing_names.add(component)
            if spi.isBulk or spi.isPrebulk or spi.well or spi.volumeDependency is not None:
                continue

            if spi.ifupdate:
                should_add = True
            else:
                should_add = sp.Name not in existing_names

            if not should_add:
                continue

            ws.append([
                date.today().isoformat(),
                sp.Name,
                sp.FinalConcentration,
                spi.finalVolume
            ])

            existing_names.add(sp.Name)

        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

        wb.save(registry_path)
        print(f"[INFO] Central registry updated → {registry_path}")

        protocol_summary = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            date_created, name, conc, vol = row
            if name in used_protocols or name in stock:
                protocol_summary.append({
                    "Date Created": date_created,
                    "Solution Name": name,
                    "Final Concentration": conc,
                    "Final Volume": vol,
                })

        return protocol_summary